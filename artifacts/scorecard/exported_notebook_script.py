#!/usr/bin/env python
# coding: utf-8

# # Phase 2B — Scorecard
# 
# **Primary scorecard**: 7-feature LR no-F6E (interpretable, less dependent on synthetic bureau).
# **Robustness scorecard** (optional, end of notebook): 22-feature LR full-F6E.
# 
# **Workflow:**
# 1. Fit WoE binning on `train_for_model` cohorts only (202509-202610)
# 2. Numeric: monotonic constraint via merge logic
# 3. Categorical: event-rate-sorted with min-bin-size merging
# 4. Freeze bins → apply to calib (202611-202612) and OOT (202701-202706)
# 5. Refit logistic regression on WoE-transformed features
# 6. Scorecard with `factor = 20/log(2)`, `base_score = 300`, `PDO = 20`, `base_odds = 1:50`
# 7. Platt calibration on calib split (NEVER OOT)
# 8. Compare: LR no-F6E raw / LR no-F6E scorecard / LightGBM primary
# 
# **Acceptance** (Gini drop scorecard vs raw LR no-F6E on OOT):
# - < 0.05 → good
# - 0.05-0.10 → acceptable with documentation
# - > 0.10 → illustrative only, not production-defensible
# 
# **Note**: optbinning 0.20.0 has a sklearn 1.8 API mismatch in this environment.
# We use a custom WoE binner in `src/scorecard.py` with the same monotonic-merge
# logic (quantile init + iterative merge until WoE is monotonic).

# In[1]:


import sys, time, json, joblib, warnings, math
from pathlib import Path

import numpy as np
import pandas as pd
import statsmodels.api as sm
warnings.filterwarnings("ignore")

REPO_ROOT = Path.cwd().parent if Path.cwd().name == "notebooks" else Path.cwd()
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from src.governance import filter_pd_eligible, get_meta_columns
from src.evaluation import (
    compute_gini, compute_ks, compute_brier, compute_calibration_metrics,
)
from src.calibration import (
    make_calibration_split, fit_platt, fit_isotonic, apply_calibrator,
)
from src.scorecard import (
    NumericBinner, CategoricalBinner, apply_woe,
    build_scorecard, score_from_woe,
)

ABT_PATH     = REPO_ROOT / "artifacts/phase2_rerun_v2/modeling_abt.parquet"
CATALOG_PATH = REPO_ROOT / "artifacts/phase2_rerun_v2/modeling_feature_catalog.csv"
OUT_DIR      = REPO_ROOT / "artifacts/scorecard"
OUT_DIR.mkdir(parents=True, exist_ok=True)

T0 = time.time()
catalog = pd.read_csv(CATALOG_PATH)
print(f"Loaded catalog: {len(catalog):,} rows")


# ## 1. Load 7-feature LR no-F6E features + temporal split

# In[2]:


# 7-feature set from Test 1 (LR no-F6E)
FEATURES_NO_F6E = [
    "app_nom_job_code",      # categorical
    "act_age_log1p",         # numeric
    "mean_cars_by_job_code", # numeric (F5A group stat)
    "app_nom_marital_status",# categorical
    "median_income_by_job_code", # numeric (F5A)
    "app_nom_branch",        # categorical
    "app_nom_city",          # categorical
]
NUMERIC_FEATS     = ["act_age_log1p", "mean_cars_by_job_code", "median_income_by_job_code"]
CATEGORICAL_FEATS = ["app_nom_job_code", "app_nom_marital_status", "app_nom_branch", "app_nom_city"]

df = pd.read_parquet(ABT_PATH, columns=FEATURES_NO_F6E + get_meta_columns())
print(f"Loaded ABT: {df.shape}")

# Temporal calibration split (same as Test 3)
TRAIN_FOR_MODEL = list(range(202509, 202611))
CALIB_PERIODS   = [202611, 202612]
df["split_new"] = make_calibration_split(df, TRAIN_FOR_MODEL, CALIB_PERIODS)
print("\nSplit counts (new temporal split):")
print(df["split_new"].value_counts().to_string())

mask_tr  = df["split_new"] == "train_for_model"
mask_cal = df["split_new"] == "calib"
mask_oot = df["split_new"] == "oot"

print("\nDR by new split:")
for s in ["train_for_model", "calib", "oot"]:
    sub = df[df["split_new"]==s]
    print(f"  {s:<18} n={len(sub):>7,}  DR={(sub['default_flag_12m']==1).mean():.4%}")


# ## 2. Fit WoE binning on `train_for_model`
# 
# - Numeric: 20 initial quantile bins → merge to enforce monotonicity + min_bin_size 5%
# - Categorical: event-rate sorted → merge to ensure each bin ≥ 5% of train rows

# In[3]:


binners = {}
y_tr = df.loc[mask_tr, "default_flag_12m"].astype(int)

for f in NUMERIC_FEATS:
    b = NumericBinner(feature=f, n_initial_bins=20, min_bin_size_frac=0.05, monotonic=True)
    b.fit(df.loc[mask_tr, f], y_tr)
    binners[f] = b
    print(f"NUM  {f:<32}  n_bins={len(b.bin_woe):>2}  IV={b.iv:.4f}  trend={b.monotonic_trend}")

for f in CATEGORICAL_FEATS:
    b = CategoricalBinner(feature=f, min_bin_size_frac=0.05)
    b.fit(df.loc[mask_tr, f], y_tr)
    binners[f] = b
    print(f"CAT  {f:<32}  n_bins={len(b.bin_woe):>2}  IV={b.iv:.4f}")


# In[4]:


# Print the binning tables
for f in FEATURES_NO_F6E:
    b = binners[f]
    print(f"\n=== {f} (IV={b.iv:.4f}) ===")
    print(f"{'bin':>3}  {'label':<35}  {'n':>7}  {'events':>6}  {'ER':>7}  {'WoE':>7}")
    for i in range(len(b.bin_woe)):
        print(f"{i:>3}  {b.get_bin_label(i):<35}  {b.bin_count[i]:>7,}  "
              f"{b.bin_events[i]:>6}  {b.bin_event_rate[i]:>6.4%}  {b.bin_woe[i]:>+7.4f}")


# ## 3. Freeze bins, apply to all splits, refit LR

# In[5]:


df_woe = apply_woe(df, binners)
woe_cols = [f"{f}_woe" for f in FEATURES_NO_F6E]
print(f"WoE columns: {woe_cols}")

# Refit LR on train_for_model
X_tr = df_woe.loc[mask_tr, woe_cols]
y_tr_arr = df_woe.loc[mask_tr, "default_flag_12m"].astype(int)
X_tr_const = sm.add_constant(X_tr, has_constant="add")
model_woe = sm.Logit(y_tr_arr, X_tr_const).fit(disp=False, maxiter=200)
print("\nWoE LR coefficients:")
for c in model_woe.params.index:
    print(f"  {c:<40}  beta={model_woe.params[c]:+.5f}  p={model_woe.pvalues[c]:.4g}")


# ## 4. Build scorecard
# 
# `factor = PDO / log(2)`, `base_score = 300`, `PDO = 20`, `base_odds = 1:50`.
# Higher score = lower default risk.

# In[6]:


PDO = 20.0
BASE_SCORE = 300.0
BASE_ODDS = 1.0 / 50.0

feature_betas = {f: model_woe.params[f"{f}_woe"] for f in FEATURES_NO_F6E}
intercept = float(model_woe.params["const"])

scorecard = build_scorecard(
    feature_betas=feature_betas,
    intercept=intercept,
    binners=binners,
    base_score=BASE_SCORE,
    pdo=PDO,
    base_odds=BASE_ODDS,
)
print(f"Scorecard rows: {len(scorecard)}")
print(scorecard.to_string(index=False))


# ## 5. Score all splits; PD via WoE-LR + Platt calibration on calib only

# In[7]:


# WoE-LR predicted probabilities on full data (all splits)
df_woe_const = sm.add_constant(df_woe[woe_cols], has_constant="add")
df["pd_woe_raw"] = model_woe.predict(df_woe_const)

# Total scorecard points per row (additive, higher = lower risk)
df["scorecard_points"] = score_from_woe(df_woe, scorecard)

# Calibrate via Platt on calib only (NEVER OOT)
y_cal = df.loc[mask_cal, "default_flag_12m"].astype(int).to_numpy()
s_cal = df.loc[mask_cal, "pd_woe_raw"].to_numpy()
platt = fit_platt(s_cal, y_cal)
df["pd_woe_platt"] = apply_calibrator(df["pd_woe_raw"].to_numpy(), platt, "platt")

print(f"Train-for-model PD raw mean    : {df.loc[mask_tr, 'pd_woe_raw'].mean():.4%}")
print(f"Train-for-model DR             : {(df.loc[mask_tr, 'default_flag_12m']==1).mean():.4%}")
print(f"Calib PD raw mean              : {df.loc[mask_cal, 'pd_woe_raw'].mean():.4%}")
print(f"OOT  PD raw mean               : {df.loc[mask_oot, 'pd_woe_raw'].mean():.4%}")
print(f"OOT  PD Platt mean             : {df.loc[mask_oot, 'pd_woe_platt'].mean():.4%}")
print(f"OOT  base rate                 : {(df.loc[mask_oot, 'default_flag_12m']==1).mean():.4%}")


# ## 6. Performance comparison on OOT

# In[8]:


def metrics_at(y, s):
    cal = compute_calibration_metrics(y, s)
    return {
        "auc"   : (compute_gini(y, s) + 1) / 2,
        "gini"  : compute_gini(y, s),
        "ks"    : compute_ks(y, s),
        "brier" : compute_brier(y, s),
        "ece"   : cal.get("ece", float("nan")),
        "mean_pred": cal.get("mean_pred", float("nan")),
        "base_rate": cal.get("base_rate", float("nan")),
    }

y_oot_arr = df.loc[mask_oot, "default_flag_12m"].astype(int).to_numpy()

scorecard_oot_raw   = metrics_at(y_oot_arr, df.loc[mask_oot, "pd_woe_raw"].to_numpy())
scorecard_oot_platt = metrics_at(y_oot_arr, df.loc[mask_oot, "pd_woe_platt"].to_numpy())

# Raw LR no-F6E baseline (Phase 2A Test 1)
with open(REPO_ROOT / "artifacts/phase2_rerun_v2/diagnostics/test1_f6e_ablation.json") as f:
    lr_nof6e = json.load(f)
lr_nof6e_oot_raw = lr_nof6e["metrics"]["oot"]

# LightGBM primary (Phase 2A retune)
with open(REPO_ROOT / "artifacts/pd_model/lightgbm_tuning_results.json") as f:
    lgb_results = json.load(f)
lgb_oot_platt = lgb_results["oot_metrics"]["pd_lgb_platt"]

# LR full-F6E + Platt (Phase 2A Test 3)
with open(REPO_ROOT / "artifacts/phase2_rerun_v2/diagnostics/test3_calibration.json") as f:
    test3 = json.load(f)
lr_full_oot_platt = test3["oot_metrics_by_calibration"]["pd_platt"]

cmp = pd.DataFrame({
    "LR_no_F6E_raw_oot":        lr_nof6e_oot_raw,
    "Scorecard_no_F6E_raw_oot": scorecard_oot_raw,
    "Scorecard_no_F6E_platt_oot": scorecard_oot_platt,
    "LR_full_F6E_platt_oot":    lr_full_oot_platt,
    "LightGBM_tuned_platt_oot": lgb_oot_platt,
})
print("=== Side-by-side OOT comparison ===")
print(cmp.round(5).to_string())

# Acceptance tier vs raw LR no-F6E
gini_raw_lr = lr_nof6e_oot_raw["gini"]
gini_drop_raw = gini_raw_lr - scorecard_oot_raw["gini"]
gini_drop_platt = gini_raw_lr - scorecard_oot_platt["gini"]
print(f"\nAcceptance test (raw LR no-F6E vs scorecard):")
print(f"  raw scorecard:   gini drop = {gini_drop_raw:+.4f}")
print(f"  Platt scorecard: gini drop = {gini_drop_platt:+.4f}")

def tier(d):
    if d < 0.05: return "good (<0.05)"
    if d < 0.10: return "acceptable (0.05-0.10)"
    return "illustrative only (>0.10)"

tier_raw = tier(gini_drop_raw)
tier_platt = tier(gini_drop_platt)
print(f"  Tier (raw):   {tier_raw}")
print(f"  Tier (Platt): {tier_platt}")


# ## 7. Save artifacts

# In[9]:


# 7a. optbin_definitions.json (binner state, JSON-friendly)
def serialize_num(b):
    return {
        "type": "numeric",
        "feature": b.feature,
        "n_initial_bins": b.n_initial_bins,
        "min_bin_size_frac": b.min_bin_size_frac,
        "monotonic": b.monotonic,
        "cuts": list(b.cuts),
        "bin_woe": list(b.bin_woe),
        "bin_event_rate": list(b.bin_event_rate),
        "bin_count": list(b.bin_count),
        "bin_events": list(b.bin_events),
        "iv": b.iv,
        "monotonic_trend": b.monotonic_trend,
    }
def serialize_cat(b):
    return {
        "type": "categorical",
        "feature": b.feature,
        "min_bin_size_frac": b.min_bin_size_frac,
        "category_to_bin": {str(k): int(v) for k, v in b.category_to_bin.items()},
        "bin_categories": [[str(c) for c in cats] for cats in b.bin_categories],
        "bin_woe": list(b.bin_woe),
        "bin_event_rate": list(b.bin_event_rate),
        "bin_count": list(b.bin_count),
        "bin_events": list(b.bin_events),
        "iv": b.iv,
    }

binner_defs = {}
for f, b in binners.items():
    binner_defs[f] = serialize_num(b) if isinstance(b, NumericBinner) else serialize_cat(b)
with open(OUT_DIR / "optbin_definitions.json", "w") as f:
    json.dump(binner_defs, f, indent=2, default=str)
print(f"Saved: {OUT_DIR / 'optbin_definitions.json'}")

# 7b. scorecard_table.csv
scorecard.to_csv(OUT_DIR / "scorecard_table.csv", index=False)
print(f"Saved: {OUT_DIR / 'scorecard_table.csv'}")

# 7c. WoE-transformed ABT (slim version, only essentials)
slim = df_woe[woe_cols + ["aid", "split", "split_new", "default_flag_12m", "fin_period"]].copy()
slim["pd_woe_raw"] = df["pd_woe_raw"]
slim["pd_woe_platt"] = df["pd_woe_platt"]
slim["scorecard_points"] = df["scorecard_points"]
slim.to_parquet(OUT_DIR / "woe_transformed_abt.parquet", index=False)
print(f"Saved: {OUT_DIR / 'woe_transformed_abt.parquet'}  ({len(slim):,} rows)")

# 7d. metrics JSON
metrics_out = {
    "splits": {
        "train_for_model_periods": TRAIN_FOR_MODEL,
        "calib_periods": CALIB_PERIODS,
        "n_train_for_model": int(mask_tr.sum()),
        "n_calib": int(mask_cal.sum()),
        "n_oot": int(mask_oot.sum()),
    },
    "features": FEATURES_NO_F6E,
    "scorecard_constants": {
        "factor": PDO / math.log(2),
        "base_score": BASE_SCORE,
        "pdo": PDO,
        "base_odds": BASE_ODDS,
    },
    "iv_per_feature": {f: float(binners[f].iv) for f in FEATURES_NO_F6E},
    "model_betas": {c: float(model_woe.params[c]) for c in model_woe.params.index},
    "model_pvalues": {c: float(model_woe.pvalues[c]) for c in model_woe.pvalues.index},
    "oot_metrics": {
        "scorecard_raw":   scorecard_oot_raw,
        "scorecard_platt": scorecard_oot_platt,
        "ref_lr_no_f6e_raw":      lr_nof6e_oot_raw,
        "ref_lr_full_f6e_platt":  lr_full_oot_platt,
        "ref_lightgbm_platt":     lgb_oot_platt,
    },
    "acceptance": {
        "gini_drop_vs_lr_no_f6e_raw": float(gini_drop_raw),
        "gini_drop_vs_lr_no_f6e_platt": float(gini_drop_platt),
        "tier_raw": tier_raw,
        "tier_platt": tier_platt,
    },
}
with open(OUT_DIR / "scorecard_metrics.json", "w") as f:
    json.dump(metrics_out, f, indent=2, default=str)
print(f"Saved: {OUT_DIR / 'scorecard_metrics.json'}")

# 7e. 1-page summary xlsx
import openpyxl
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "Summary"
rows = [
    ["Phase 2B Scorecard Summary", ""],
    ["", ""],
    ["Features", ", ".join(FEATURES_NO_F6E)],
    ["n train_for_model", int(mask_tr.sum())],
    ["n calib", int(mask_cal.sum())],
    ["n oot", int(mask_oot.sum())],
    ["", ""],
    ["IV (sum)", round(sum(binner_defs[f]['iv'] for f in FEATURES_NO_F6E), 4)],
    ["Acceptance tier (Platt)", tier_platt],
    ["Gini drop vs raw LR (Platt)", round(gini_drop_platt, 4)],
    ["", ""],
    ["OOT metric", "Scorecard Platt"],
    ["AUC", round(scorecard_oot_platt['auc'], 4)],
    ["Gini", round(scorecard_oot_platt['gini'], 4)],
    ["KS", round(scorecard_oot_platt['ks'], 4)],
    ["Brier", round(scorecard_oot_platt['brier'], 5)],
    ["ECE", round(scorecard_oot_platt['ece'], 5)],
    ["Mean predicted", round(scorecard_oot_platt['mean_pred'], 5)],
    ["Base rate", round(scorecard_oot_platt['base_rate'], 5)],
]
for r in rows:
    ws.append(r)
ws2 = wb.create_sheet("Comparison")
ws2.append(["model", "AUC", "Gini", "KS", "Brier", "ECE", "mean_pred"])
for name, m in [
    ("LR no-F6E raw",         lr_nof6e_oot_raw),
    ("Scorecard no-F6E raw",  scorecard_oot_raw),
    ("Scorecard no-F6E Platt",scorecard_oot_platt),
    ("LR full-F6E Platt",     lr_full_oot_platt),
    ("LightGBM tuned Platt",  lgb_oot_platt),
]:
    ws2.append([name,
                round(m['auc'], 4),
                round(m['gini'], 4),
                round(m.get('ks', 0), 4),
                round(m['brier'], 5),
                round(m.get('ece', 0), 5),
                round(m.get('mean_pred', 0), 5)])
ws3 = wb.create_sheet("Scorecard")
ws3.append(list(scorecard.columns))
for _, row in scorecard.iterrows():
    ws3.append([row[c] for c in scorecard.columns])
wb.save(OUT_DIR / "scorecard_summary.xlsx")
print(f"Saved: {OUT_DIR / 'scorecard_summary.xlsx'}")
print(f"\nWall (Phase 2B): {time.time()-T0:.1f}s")


# ## 8. (OPTIONAL) Robustness scorecard — 22-feature full-F6E
# 
# Only run if Section 1-7 succeeded. Same WoE workflow on the LR full-F6E final
# features. Saved separately to `artifacts/scorecard/robustness_full_f6e/`.

# In[10]:


import time
T1 = time.time()
ROB_DIR = OUT_DIR / "robustness_full_f6e"
ROB_DIR.mkdir(parents=True, exist_ok=True)

# Load 22-feature set from Phase 2A
final_features_full = pd.read_csv(REPO_ROOT / "artifacts/phase2_rerun_v2/final_feature_set.csv")["feature"].tolist()
print(f"Robustness: 22-feature LR full-F6E features:")
for f in final_features_full:
    print(f"  - {f}")

# Determine numeric vs categorical via catalog
cat_idx = catalog.set_index("feature_name")
KNOWN_CAT = {"app_nom_branch","app_nom_gender","app_nom_job_code",
             "app_nom_marital_status","app_nom_city","app_nom_home_status"}
ROB_NUMERIC = []; ROB_CAT = []
for f in final_features_full:
    if f in KNOWN_CAT:
        ROB_CAT.append(f)
    else:
        ROB_NUMERIC.append(f)
print(f"\nNumeric: {len(ROB_NUMERIC)}, Categorical: {len(ROB_CAT)}")


# In[11]:


# Load all 22 features + meta
df_rob = pd.read_parquet(ABT_PATH, columns=final_features_full + get_meta_columns())
df_rob["split_new"] = make_calibration_split(df_rob, TRAIN_FOR_MODEL, CALIB_PERIODS)
mask_tr2  = df_rob["split_new"] == "train_for_model"
mask_cal2 = df_rob["split_new"] == "calib"
mask_oot2 = df_rob["split_new"] == "oot"
y_tr2 = df_rob.loc[mask_tr2, "default_flag_12m"].astype(int)

binners_rob = {}
for f in ROB_NUMERIC:
    b = NumericBinner(feature=f, n_initial_bins=20, min_bin_size_frac=0.05, monotonic=True)
    b.fit(df_rob.loc[mask_tr2, f], y_tr2)
    binners_rob[f] = b
for f in ROB_CAT:
    b = CategoricalBinner(feature=f, min_bin_size_frac=0.05)
    b.fit(df_rob.loc[mask_tr2, f], y_tr2)
    binners_rob[f] = b
print("Per-feature IV (robustness):")
for f in final_features_full:
    print(f"  {f:<45}  IV={binners_rob[f].iv:.4f}  bins={len(binners_rob[f].bin_woe)}")


# In[12]:


df_rob_woe = apply_woe(df_rob, binners_rob)
woe_cols_rob = [f"{f}_woe" for f in final_features_full]
X_tr_r = df_rob_woe.loc[mask_tr2, woe_cols_rob]
y_tr_r = df_rob_woe.loc[mask_tr2, "default_flag_12m"].astype(int)
X_tr_r_const = sm.add_constant(X_tr_r, has_constant="add")
model_rob = sm.Logit(y_tr_r, X_tr_r_const).fit(disp=False, maxiter=200)
print(f"Robustness LR refit: {len(final_features_full)} features")
# Show only significant coefficients
sig = model_rob.pvalues[model_rob.pvalues < 0.05]
print(f"Significant (p<0.05): {len(sig)} of {len(model_rob.pvalues)}")

feature_betas_rob = {f: float(model_rob.params[f"{f}_woe"]) for f in final_features_full}
intercept_rob = float(model_rob.params["const"])

scorecard_rob = build_scorecard(
    feature_betas=feature_betas_rob,
    intercept=intercept_rob,
    binners=binners_rob,
    base_score=BASE_SCORE,
    pdo=PDO,
    base_odds=BASE_ODDS,
)
print(f"Robustness scorecard rows: {len(scorecard_rob)}")


# In[13]:


# Score and Platt-calibrate
df_rob_woe_const = sm.add_constant(df_rob_woe[woe_cols_rob], has_constant="add")
df_rob["pd_rob_raw"] = model_rob.predict(df_rob_woe_const)
y_cal2 = df_rob.loc[mask_cal2, "default_flag_12m"].astype(int).to_numpy()
s_cal2 = df_rob.loc[mask_cal2, "pd_rob_raw"].to_numpy()
platt_rob = fit_platt(s_cal2, y_cal2)
df_rob["pd_rob_platt"] = apply_calibrator(df_rob["pd_rob_raw"].to_numpy(), platt_rob, "platt")

y_oot2 = df_rob.loc[mask_oot2, "default_flag_12m"].astype(int).to_numpy()
rob_raw   = metrics_at(y_oot2, df_rob.loc[mask_oot2, "pd_rob_raw"].to_numpy())
rob_platt = metrics_at(y_oot2, df_rob.loc[mask_oot2, "pd_rob_platt"].to_numpy())

print("Robustness scorecard OOT metrics:")
print(pd.DataFrame({"raw": rob_raw, "platt": rob_platt}).round(5).to_string())

# Save robustness artifacts
scorecard_rob.to_csv(ROB_DIR / "scorecard_table.csv", index=False)
with open(ROB_DIR / "scorecard_metrics.json", "w") as f:
    json.dump({
        "n_features": len(final_features_full),
        "features": final_features_full,
        "iv_per_feature": {f: float(binners_rob[f].iv) for f in final_features_full},
        "betas": feature_betas_rob,
        "intercept": intercept_rob,
        "oot_metrics": {"raw": rob_raw, "platt": rob_platt},
    }, f, indent=2, default=str)
print(f"\nSaved: {ROB_DIR / 'scorecard_table.csv'}")
print(f"Saved: {ROB_DIR / 'scorecard_metrics.json'}")
print(f"Robustness wall: {time.time()-T1:.1f}s")
print(f"\nTotal Phase 2B wall: {time.time()-T0:.1f}s")


# ## 9. Done
# 
# Primary scorecard (7-feature LR no-F6E) and optional robustness scorecard
# (22-feature LR full-F6E) both built. See `artifacts/scorecard/` for outputs.
